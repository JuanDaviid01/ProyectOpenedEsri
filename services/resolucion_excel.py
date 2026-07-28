from pathlib import Path
import os
# pyrefly: ignore [missing-import]
from openpyxl import load_workbook
from datetime import datetime

def obtener_ruta_base():
    # Sube un nivel adicional para ir del directorio services/ a la raíz del proyecto
    return Path(__file__).resolve().parent.parent

def obtener_ruta_excel_resoluciones():
    ruta = obtener_ruta_base() / "documentos" / "FCI & RESOLUCIONES 2025.xlsx"
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel de resoluciones: {ruta}")
    return ruta

def obtener_y_registrar_resolucion(datos, usuario="PAOLA"):
    """
    Toma el último consecutivo real del histórico del Excel,
    genera el siguiente número y registra una nueva fila.

    Columnas esperadas:
    A = fecha
    B = número resolución
    C = usuario
    D = descripción
    """
    ruta_excel = obtener_ruta_excel_resoluciones()

    wb = load_workbook(ruta_excel)
    ws = wb.active

    ultima_fila = ws.max_row

    # Buscar hacia arriba la última fila con número en columna B
    ultimo_numero = None
    for fila in range(ultima_fila, 0, -1):
        valor = ws[f"B{fila}"].value
        if valor is not None and str(valor).strip() != "":
            try:
                ultimo_numero = int(valor)
                break
            except ValueError:
                continue

    if ultimo_numero is None:
        ultimo_numero = 1000

    nuevo_numero = ultimo_numero + 1
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")

    nombre_estudiante = str(datos.get("nombre_estudiante") or "").strip()
    programa = str(datos.get("programa") or "").strip()

    if not nombre_estudiante or nombre_estudiante.upper() == "N/A":
        nombre_estudiante = "SIN NOMBRE"

    descripcion = f"Res. {nuevo_numero} Homologación {programa} {nombre_estudiante}".strip()

    # Agregar nueva fila al final
    ws.append([
        fecha_hoy,        # Columna A
        nuevo_numero,     # Columna B
        usuario,          # Columna C
        descripcion       # Columna D
    ])

    wb.save(ruta_excel)

    return str(nuevo_numero)
